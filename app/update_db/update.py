import abc
import argparse
from collections import defaultdict
import logging
from functools import wraps
import typing as T
from typing import (
    Type, Tuple, List, Dict, Union, Callable, Iterator,
    Literal, Any)

from openpyxl import load_workbook

from ..models import (
    UNKNOWN_SYNT_FUNCTION_OF_ANCHOR,
    SYNT_FUNCTION_OF_ANCHOR_VALUES,
    Construction,
    ConstructionVariant,
    GeneralInfo,
    Change,
    Constraint,
    FormulaElement
)


logging.basicConfig(handlers=(logging.StreamHandler(),))
print(__name__)
logger = logging.getLogger(__name__)


ORIG_SHEET_NAME2DB_TABLE_NAME = {
    "cnstruct": "construction",
    "gen_inf": "general_info",
    "ch": "changes",
    "cnstraint": "constraints",
}


SHEET_TO_CLASS = {
    "construction": Construction,
    "general_info": GeneralInfo,
    "changes": Change,
    "constraints": Constraint,
}


class StrLoweringDict(dict):
    def get(self, item: str, default: Any = None):
        # logger.debug(f"item requested: `{item}`")
        # if default is None:
        default = item.lower()
        return super(StrLoweringDict, self).get(item, default)

    def __repr__(self):
        return f"{self.__class__.__name__}({super().__repr__()})"


default_colname_correction = StrLoweringDict()


SHEET2COLUMN2CORRECTION = {
    "general_info": StrLoweringDict({
        "construction_name": "name",
    }),
    "construction": StrLoweringDict({
        "construction_id": "id",
        "in_russian_constructicon": "in_rus_constructicon",
        "number_in_russian_constructicon": "rus_constructicon_id",
        "constructicon_morphosyntags": "morphosyntags",
        "constructicon_semantags": "semantags",
    }),
    "constraints": StrLoweringDict({
        # "constraint_id": "id",
        "syntactic_constraints": "syntactic",
        "semantic_constraints": "semantic",
    }),
    "changes": StrLoweringDict({
        "change_id": "id",
        "part_of_construction_changed": "stage",
        "first_entry": "first_attested",
        "last_entry": "last_attested",
        "frequency_(trend)": "frequency_trend",
    }),
}


SHEET2MUSTHAVE_COLUMNS = {
    "general_info": {"construction_id"},
    "changes": {"construction_id", "stage", "level"}
}


SHEET2DISCARDED_COLUMNS = {
    "constraints": {"constraint_id"}
}


CONSTRUCTION_ID2VARIANTSMET = {}
# CONSTRUCTION_IDS 


class EOF(object): ...

VARIANTS_SEPS: T.Tuple[str, str] = ("/", "|")


def read_until(
    formula_it: Iterator, finish_at_predicate: Callable[[str], bool]
) -> Tuple[str, Union[str, EOF]]:
    res = []
    while True:
        next_ = next(formula_it, EOF)
        if finish_at_predicate(next_):
            return "".join(res), next_
        res.append(next_)


# TODO: parse alternatives
def tokenize_formula(
    formula: str, elems_sep: str = " ",
    span_start: str = "(", span_end: str = ")",
    variants_seps: T.Tuple[str, ...] = VARIANTS_SEPS, 
    are_span_symbs_optionality_symb=True,
) -> List[Dict]:
    """Tokenizes formula into word or span"""
    SPECIAL = [elems_sep, span_start, span_end, *variants_seps, EOF]
    # SPECIAL = [elems_sep, span_start, span_end, EOF]

    parts = cur_part = []
    queue = [parts]

    formula_it = iter(formula)
    symbol = ""

    while True:
        symbol, prev = next(formula_it, EOF), symbol
        # is_span_open = False

        # print("after read", symbol, prev, cur_part, parts, queue, sep="\n  ")

        if symbol not in SPECIAL:
            cur_element, special = read_until(formula_it, lambda symbol: symbol in SPECIAL)

            cur_part.append({"val": symbol + cur_element})
            symbol, prev = special, symbol

        if symbol == span_start:
            cur_part = []
            # is_span_open = True
            queue.append(cur_part)
        elif symbol == span_end:
            # print("span_end before append", symbol, prev, cur_part, parts, queue, sep="\n  ")
            # parts.append({"type": "maybe_span", "parts": queue.pop()})
            cur_result = {"type": "maybe_span", "val": queue.pop()}
            # is_span_open = False
            cur_part = queue[-1]
            cur_part.append(cur_result)
            # print("span_end after append and cur_part", symbol, prev, cur_part, parts, queue, sep="\n  ")
        # elif symbol in variants_seps:
        #     is_cur_result_type_span = len(queue[-1]) > 1
            
        #     cur_result = {"val": queue.pop()}
        #     if is_cur_result_type_span:
        #         cur_result["type"] = "maybe_span"
            
            
        elif symbol is EOF:
            # print("EOF", symbol, prev, cur_part, parts, queue, sep="\n  ")
            break

    return parts


def find_alternatives(
    val: str, variants_seps: T.Tuple[str, ...] = VARIANTS_SEPS
) -> T.Optional[T.List[str]]:
    for sep in variants_seps:
        if sep in val:
            return val.split(sep)
    return None


def flatten_span(
    span_token_values: List[Dict[str, Any]], depth=1, order=0
) -> List[Dict[str, Union[str, int]]]:
    flattened_token_list = []

    if len(span_token_values) == 1:
        tok_desc = {"value": span_token_values[0]["val"], "is_optional": True}
        if depth > 1:
            tok_desc["depth"] = depth - 1
        return [tok_desc]
    
    for tok in span_token_values:
        if tok.get("type") == "maybe_span":
            flattened_token_list.extend(flatten_span(tok["val"], depth=depth + 1))
        else:
            flattened_token_list.append({"value": tok["val"], "depth": depth,
                                         "is_optional": False})

    return flattened_token_list


def parse_formula(
    formula: str, el_variants_sep="/",
) -> List[Dict]:
    elements = []
    order = 0

    tokens = tokenize_formula(formula)
    for token in tokens:
        if token.get("type") == "maybe_span":
            span_elements = flatten_span(token["val"])

            for _order, span_element in enumerate(span_elements, order):
                span_element["order"] = _order
            order = _order + 1

            elements.extend(span_elements)
        else:
            elements.append({"value": token["val"], "order": order})
            order += 1

    return elements


# TODO: special processing for NP (=NP-nom) ?
def parse_formula_old(
        formula: str,
        el_variants_sep="/",
) -> List[Dict]:
    data = []

    # TODO: improve tokenization to take spans into account
    for i, element in enumerate(formula.split(" ")):
        self_result_elements = []

        has_brackets = any(br in element for br in ("(", ")"))
        if has_brackets:
            element = element.strip("()")

        has_variants = el_variants_sep in element
        if has_variants:
            actual_elements = element.split(el_variants_sep)
            first_element_data = {
                "value": actual_elements[0], "order": i,
                "is_optional": has_brackets, "has_variants": has_variants
            }
            self_result_elements.append(first_element_data)
            self_result_elements.extend([
                {"value": variant_el, "order": i, "is_optional": has_brackets}
                for variant_el in actual_elements[1:]
            ])
        else:
            self_result_elements = [
                {"value": element, "order": i, "is_optional": has_brackets}]

        # TODO: special parsing / saving of nom nps (simply `n` / `np`)

        data.extend(self_result_elements)

    return data


def fix_construction_id(
    construction_id: str, variant_sep="-",
    corrections=str.maketrans({".": "0", "?": "9", " ": None})
):
    logger.debug(f"constr id: {construction_id}, type is {type(construction_id)}")
    if isinstance(construction_id, (int, float)):
        return int(construction_id)

    if variant_sep in construction_id:
        main_id_part, variant_id, *extra = construction_id.split(variant_sep)
        if extra:
            raise ValueError(f"construction id: extra variant_sep in `{main_id_part}`")
    else:
        main_id_part, variant_id = construction_id, ""

    variants_met = CONSTRUCTION_ID2VARIANTSMET

    # variant_id, prev_variant_id = len(variants_met), variant_id
    # variants_met.append(variant_id)
    # logger.debug(f"previous variant_id:  {prev_variant_id}, new: {variant_id}")


    main_id_part = main_id_part.translate(corrections)
    logger.debug(f"old: {construction_id}, new main part: {main_id_part}")

    if "(" in main_id_part and ")" in main_id_part:
        num = main_id_part[:main_id_part.index("(")]
        group = main_id_part[main_id_part.index("(")+1:main_id_part.index(")")]

        return int(group + num + str(variant_id))

    else:
        print(main_id_part)
        raise ValueError
    


def fix_construction_id_wrapper(phrase_dict, id_key="construction_id"):
    orig_id = phrase_dict[id_key]
    maybe_id = fix_construction_id(phrase_dict[id_key])
    print("constr ids:", orig_id, maybe_id)
    if str(orig_id) == str(maybe_id):
        maybe_id = int(str(maybe_id) + phrase_dict.get("group_number", "000"))
    phrase_dict[id_key] == maybe_id


id_to_change_object = {}


class ChangesManager:
    def __init__(self) -> None:
        self.ids_construction2changes: T.Dict[int, T.List[int]] = {}

    def add_construction(self, id_: int, obj: Construction):
        self.ids_construction2changes.setdefault(id_, [])
    
    def add_change(self, id_: int, construction_id: int):
        self.ids_construction2changes.setdefault(construction_id, []).append(id_)



def fix_values(
    phrase_dict: Dict[str, Union[str, int, None]],
    model_class: Type[Union[Construction, GeneralInfo, Change, Constraint]]
):
    logger.debug(f"fixing values: {phrase_dict}, model: {model_class}")
    if model_class is Construction:
        phrase_dict["orig_id"] = phrase_dict["id"]
        # fix_construction_id_wrapper(phrase_dict, id_key="id")
        phrase_dict["id"] = fix_construction_id(phrase_dict["id"])
    elif "construction_id" in phrase_dict:
        phrase_dict["construction_id"] = fix_construction_id(phrase_dict["construction_id"])

    if model_class is Change:
        former_change = phrase_dict.pop("former_change")
        if former_change:
            if not isinstance(former_change, int):
                previous_changes_ids = [int(_id.strip()) for _id in former_change.split(",")]
            else:
                previous_changes_ids = [former_change]
            for _id in previous_changes_ids:
                obj = id_to_change_object.get(_id)
                if obj:
                    phrase_dict.setdefault("previous_changes", []).append(obj)
                else:
                    raise ValueError(f"no change defined: {_id} (from {phrase_dict['id']})")
        else:
            phrase_dict["previous_changes"] = []

    return phrase_dict


def get_formula_element_variants(variation_val: str) -> Dict[str, List[str]]:
    elem2variants = {}

    for single_variation_desc in variation_val.split("\n"):
        if not (":" in single_variation_desc and "," in single_variation_desc):
            continue
        if (":" in single_variation_desc) ^ ("," in single_variation_desc):
            # there may be error in markup
            continue

        described_formula_el, el_variants_str = single_variation_desc.split(":")
        el_variants = [el_var.strip() for el_var in el_variants_str.split(",")]

        elem2variants[described_formula_el] = el_variants

    return elem2variants


def to_formula(
    formula_s: str, element_model = FormulaElement,
    formula_parser: Callable[[str], List[Dict[str, str]]]=parse_formula
) -> None:
    formula_elements = formula_parser(formula_s)
    formula_elements_vals = [element_model(**el_data)
                             for el_data in formula_elements]
    return formula_elements_vals


def process_construction(
    phrase_dict: Dict[str, Union[str, int]], constr: Construction,
    formula_parser: Callable[[str], List[Dict[str, str]]] = parse_formula,
    verbose=False
) -> None:
    """Parse and add to construction object variants and their formula elements

    :param phrase_dict:
    :param constr:
    :param formula_parser:
    :param verbose:
    :return:
    """

    synt_function = phrase_dict.get("synt_function_of_anchor")
    logger.debug(f"function is: {synt_function}")
    if synt_function not in SYNT_FUNCTION_OF_ANCHOR_VALUES:
        if synt_function is None:
            first_synt_function = UNKNOWN_SYNT_FUNCTION_OF_ANCHOR
        else:
            first_synt_function = synt_function.split()[0]
            if first_synt_function not in SYNT_FUNCTION_OF_ANCHOR_VALUES:
                first_synt_function = UNKNOWN_SYNT_FUNCTION_OF_ANCHOR
        phrase_dict["synt_function_of_anchor"] = first_synt_function
        constr.synt_function_of_anchor = first_synt_function
        logger.debug(f"final function is: {first_synt_function}")

    formula_elements = formula_parser(phrase_dict["formula"])
    formula_elements_vals = [FormulaElement(**el_data)
                             for el_data in formula_elements]
    if verbose:
        print(FormulaElement, formula_elements_vals)

    constr.formula_elements.extend(formula_elements_vals)

    main_constr = ConstructionVariant(is_main=True)
    main_constr.formula_elements.extend(formula_elements_vals)
    construction_variants_vals = [main_constr]

    construction_variants = [
        var for var in (phrase_dict["variation"] or "").split("\n")
        if var and not any(symb in var for symb in ":,")
    ]
    print(construction_variants)

    for variant in construction_variants:
        constr_variant = ConstructionVariant(formula=variant)
        variant_formula_elements = formula_parser(variant)
        variant_formula_elements_vals = [FormulaElement(**el_data)
                                         for el_data in variant_formula_elements]
        constr_variant.formula_elements.extend(variant_formula_elements_vals)
        construction_variants_vals.append(constr_variant)

    if verbose:
        print(ConstructionVariant, construction_variants_vals)

    constr.variants.extend(construction_variants_vals)


def process_change(
    phrase_dict: Dict[str, Union[str, int]], change: Change,
    formula_parser: Callable[[str], List[Dict[str, str]]] = parse_formula,
    verbose=False
) -> None:
    """Parse stage formula and add it to construction variants"""
    main_stage_variant = ConstructionVariant(
        construction_id=change.construction_id, is_main=True
    )

    stage = phrase_dict["stage"]
    main_stage_variant.formula = stage
    if stage not in {"entire_construction"}:
        elements = to_formula(stage)
        main_stage_variant.formula_elements = elements

    change.variants.append(main_stage_variant)


extra_processing = {
    Construction: process_construction,
    Change: process_change,
}


RowDict = T.Dict[str, T.Any]


def not_empty(collection: T.Iterable):
    return bool(collection)

def make_has_dict_keys(keys: T.List[str]) -> T.Callable[[RowDict], bool]:
    def has_dict_keys(d: RowDict) -> str:
        return all(key in d for key in keys)
    return has_dict_keys

class BaseInputTable(abc.ABC):
    MUST_HAVE_COLS = []
    row_filters = [not_empty]

    def __init__(self) -> None:
        self.row_filters = self.row_filters + [
            make_has_dict_keys(self.MUST_HAVE_COLS)
        ]

    def is_row_okay(self, row: RowDict) -> bool:
        return all(filt(row) for filt in self.row_filters)
    
    def process_row(self, row: RowDict) -> RowDict:
        return row


class ConstructionInputTable(BaseInputTable):
    MUST_HAVE_COLS = []



def convert_column_title(orig_title: str) -> str:
    return orig_title.replace(" ", "_").lower()


def parse(filename: str, use_old_sheet_names=True, verbose=False):
    wb = load_workbook(filename, data_only=True)

    # parse.idscorrection = {}

    data = []
    print(wb.worksheets)

    for sheet in wb.worksheets:
        if not use_old_sheet_names:
            corrected_sheet_name = ORIG_SHEET_NAME2DB_TABLE_NAME.get(sheet.title)
        else:
            corrected_sheet_name = sheet.title
        model_class = SHEET_TO_CLASS.get(corrected_sheet_name)
        if model_class is None:
            continue

        rows_iter = sheet.iter_rows()

        # TODO: formula versus value
        title_row = next(rows_iter)
        title_row_values = [convert_column_title(cell.value)
                            for cell in title_row if cell.value]
        this_sheet_corrections = SHEET2COLUMN2CORRECTION.get(corrected_sheet_name, {})
        logger.debug(f"this sheet corrections: {this_sheet_corrections}")

        title_row_values = [
            this_sheet_corrections.get(value, value)
            for value in title_row_values
        ]
        print("title is", sheet.title, model_class, corrected_sheet_name,
              this_sheet_corrections.get("change_id"),
              title_row_values)

        for row in rows_iter:
            cell_values = [cell.value.strip() if isinstance(cell.value, str) else cell.value
                           for cell in row]
            if not any(cell_values):
                continue

            phrase_dict = dict(zip(title_row_values, cell_values))

            discard_row = False
            for col in SHEET2MUSTHAVE_COLUMNS.get(corrected_sheet_name, ()):
                if not phrase_dict.get(col):
                    discard_row = True
                    break

            if model_class is Construction:
                id_ = phrase_dict["id"]
                if id_ in CONSTRUCTION_ID2VARIANTSMET:
                    discard_row = True
                else:
                    CONSTRUCTION_ID2VARIANTSMET[id_] = True

            if discard_row:
                continue

            try:
                fix_values(phrase_dict, model_class)

                if verbose:
                    print(model_class, phrase_dict)
                logger.debug(f"about to add to `{model_class}` this: {phrase_dict}")

                values = model_class(**phrase_dict)
                if model_class is Change:
                    id_ = phrase_dict["id"]
                    if not id_ in id_to_change_object:
                        id_to_change_object[id_] = values
                    else:
                        raise ValueError(f"repeating change_id: {id_}")

                if model_class in extra_processing:
                    extra_processing[model_class](phrase_dict, values, verbose=verbose)

                data.append(values)
            
            except ValueError as e:
                logger.warning(f"couldn't add {phrase_dict}: {e}")

    if verbose:
        for item in data:
            print(item, end="\n\n")

    return data


if __name__ == "__main__":
    from ..database_utils import init_db, make_database

    parser = argparse.ArgumentParser(
        description="Append data from a suitable .xlsx (.README) to"
                    "an (existing) database that is configured for this app")

    parser.add_argument("file", metavar="F", type=str,
                        help="an existing database file path")
    parser.add_argument("--database-url", type=str, default=None,
                        help="an optional sqlalchemy uri to use a different database")
    # parser.add_argument("--database-name", type=str, default=None,
    #                     help="an optional name for the database with same scheme")
    parser.add_argument("-d", "--old", action="store_true",
                        help="use olD sheet names")
    parser.add_argument("-i", "--init", action="store_true",
                        help="whether to initialize user database")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="whether to print values as they are processed")
    parser.add_argument("-q", "--sql-quiet", action="store_true",
                        help="whether to silence SQL console logging")
    args = parser.parse_args()

    kwargs = (dict(sqlalchemy_echo=False, sqlalchemy_echo_pool=False)
              if args.sql_quiet else {})

    if args.database_url is None:
        from ..database import engine, db_session, Base
    else:
        engine, db_session, Base = make_database(args.database_url, **kwargs)

    if args.init:
        init_db(Base, engine)

    data = parse(args.file, use_old_sheet_names=args.old, verbose=args.verbose)
    print(len(data))

    # db_session.add_all(data)
    for piece in data:
        db_session.add(piece)
        db_session.commit()

    # db_session.commit()

    print(f"Commit made!")
